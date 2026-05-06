import requests
import openpyxl
from io import BytesIO

# Create a test Excel file with the user's actual column headers
workbook = openpyxl.Workbook()
sheet = workbook.active

# Add the user's exact column headers
headers = [
    "Days to Maturity", "Years to Maturity", "Rate", "Nominal", "Carrying Value", 
    "Maturity Value", "Discount Rate", "Present Value", "Impairment (US$)", "Impairment ZWL$",
    "Fund Name", "NSS rates", "Discount rates", "Time value", "Present Value", 
    "Time value impairment", "Credit Impairment", "total Impairment (US$)", "total Impairment ZWL$",
    "Portfolio", "Short Name", "Counterparty", "Maturity Date", "Issue Date", 
    "Valuation date", "Tenure"
]

# Add headers to first row
for i, header in enumerate(headers, 1):
    sheet.cell(row=1, column=i, value=header)

# Add sample data for 15 rows to test pagination
for row in range(2, 17):  # 15 rows of data
    sheet.cell(row=row, column=1, value=f"Day {row-1}")
    sheet.cell(row=row, column=2, value=f"{row-1} years")
    sheet.cell(row=row, column=3, value=f"{4.5 + (row-2)*0.1}%")
    sheet.cell(row=row, column=4, value=f"${1000 + (row-2)*100}")
    sheet.cell(row=row, column=5, value=f"${950 + (row-2)*50}")
    sheet.cell(row=row, column=6, value=f"${1000 + (row-2)*100}")
    sheet.cell(row=row, column=7, value=f"{0.05 + (row-2)*0.01}")
    sheet.cell(row=row, column=8, value=f"${900 + (row-2)*80}")
    sheet.cell(row=row, column=9, value=f"${10 + (row-2)*5}")
    sheet.cell(row=row, column=10, value=f"${100 + (row-2)*50}")
    sheet.cell(row=row, column=11, value=f"Fund_{row-1}")
    sheet.cell(row=row, column=12, value=f"{3.5 + (row-2)*0.05}%")
    sheet.cell(row=row, column=13, value=f"{0.04 + (row-2)*0.01}")
    sheet.cell(row=row, column=14, value=f"${50 + (row-2)*10}")
    sheet.cell(row=row, column=15, value=f"${850 + (row-2)*70}")
    sheet.cell(row=row, column=16, value=f"${5 + (row-2)*2}")
    sheet.cell(row=row, column=17, value=f"${2 + (row-2)*1}")
    sheet.cell(row=row, column=18, value=f"${15 + (row-2)*7}")
    sheet.cell(row=row, column=19, value=f"${150 + (row-2)*70}")
    sheet.cell(row=row, column=20, value=f"Portfolio_{row-1}")
    sheet.cell(row=row, column=21, value=f"Short_{row-1}")
    sheet.cell(row=row, column=22, value=f"Counterparty_{row-1}")
    sheet.cell(row=row, column=23, value=f"2024-{row-1:02d}-01")
    sheet.cell(row=row, column=24, value=f"2023-{row-1:02d}-01")
    sheet.cell(row=row, column=25, value=f"2024-04-28")
    sheet.cell(row=row, column=26, value=f"{row-1} years")

# Save to BytesIO
excel_buffer = BytesIO()
workbook.save(excel_buffer)
excel_buffer.seek(0)

# Test upload
url = "http://localhost:5000/api/upload"
files = {
    'file': ('user_financial_data.xlsx', excel_buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
}
data = {
    'instrument_type': 'treasury_bills'
}

try:
    response = requests.post(url, files=files, data=data)
    print(f"Status Code: {response.status_code}")
    
    if response.status_code == 200:
        result = response.json()
        print(f"Success: {result.get('success')}")
        print(f"Data rows: {len(result.get('data', {}).get('data', []))}")
        
        # Show column headers
        display_headers = result.get('data', {}).get('display_headers', [])
        print(f"\nColumn Headers ({len(display_headers)}):")
        for i, header in enumerate(display_headers, 1):
            print(f"  {i:2d}. {header}")
        
        # Show first 3 rows of data
        data_rows = result.get('data', {}).get('data', [])
        if data_rows:
            print(f"\nFirst 3 rows of data:")
            for i, row in enumerate(data_rows[:3]):
                print(f"Row {i+1}: {dict(list(row.items())[:5])}...")  # Show first 5 columns
    else:
        print(f"Error: {response.text}")
        
except Exception as e:
    print(f"Error: {e}")
