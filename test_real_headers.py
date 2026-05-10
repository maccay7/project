import requests
import openpyxl
from io import BytesIO
import os

def create_test_excel():
    """Create test Excel file from environment data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Get headers from environment
    headers_env = os.environ.get('EXCEL_HEADERS', '')
    if headers_env:
        headers = headers_env.split(',')
    else:
        print("ERROR: EXCEL_HEADERS environment variable not set")
        return None
    
    # Add headers
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    
    # Get row count from environment
    rows = int(os.environ.get('TEST_ROWS', '0'))
    if rows == 0:
        print("ERROR: TEST_ROWS environment variable not set")
        return None
    
    # Add data rows
    for i in range(2, rows + 2):
        for col in range(1, len(headers) + 1):
            ws.cell(i, col, f"value_{i-1}_{col}")
    
    return wb

# Get API URL from environment
api_url = os.environ.get('API_URL', 'http://localhost:5000/api/upload')
instrument_type = os.environ.get('INSTRUMENT_TYPE', '')

if not instrument_type:
    print("ERROR: INSTRUMENT_TYPE environment variable not set")
    exit(1)

# Create and upload
wb = create_test_excel()
if not wb:
    exit(1)

buffer = BytesIO()
wb.save(buffer)
buffer.seek(0)

try:
    response = requests.post(
        api_url,
        files={'file': ('data.xlsx', buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
        data={'instrument_type': instrument_type}
    )
    
    print(f"Status: {response.status_code}")
    if response.status_code == 200:
        data = response.json().get('data', {})
        rows = data.get('data', [])
        print(f"Rows: {len(rows)}")
    else:
        print(f"Error: {response.text}")
        
except Exception as e:
    print(f"Error: {e}")