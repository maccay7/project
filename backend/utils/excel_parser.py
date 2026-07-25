import pandas as pd
import openpyxl
import io
import os
from pathlib import Path

def parse_full_workbook(file_path_or_bytes, instrument_type='money-market', max_rows=100000):
    """
    Parse an Excel file and return all sheets with data, headers, fullData, and merged ranges.
    
    Args:
        file_path_or_bytes: file path (str/Path) or bytes of the file
        instrument_type: unused but kept for compatibility
        max_rows: maximum number of rows to read per sheet (default 100,000)
    
    Returns:
        dict with keys: 'success' (bool), 'sheets' (list), 'error' (str if failed)
    """
    try:
        if isinstance(file_path_or_bytes, (str, Path)):
            if os.path.isfile(file_path_or_bytes):
                file_path = file_path_or_bytes
                wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
                sheets = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    full_data = []
                    max_col = 0
                    # Read rows up to max_rows
                    for row in ws.iter_rows(values_only=True, max_row=max_rows):
                        if not any(cell is not None for cell in row):
                            continue
                        row_list = [cell if cell is not None else '' for cell in row]
                        full_data.append(row_list)
                        if len(row_list) > max_col:
                            max_col = len(row_list)
                    # Pad rows to have consistent column count
                    for row in full_data:
                        if len(row) < max_col:
                            row.extend([''] * (max_col - len(row)))

                    # Extract merged ranges (0‑based indices for frontend)
                    merged_ranges = []
                    for merge in ws.merged_cells.ranges:
                        merged_ranges.append({
                            'min_row': merge.min_row - 1,
                            'max_row': merge.max_row - 1,
                            'min_col': merge.min_col - 1,
                            'max_col': merge.max_col - 1
                        })

                    # Build headers and data dicts (for compatibility)
                    if full_data:
                        first_row = full_data[0]
                        is_header_row = any(isinstance(cell, str) and not cell.isdigit() for cell in first_row)
                        if is_header_row:
                            headers = [str(h) for h in first_row]
                            data_rows = full_data[1:]
                        else:
                            headers = [f"Column_{i+1}" for i in range(len(first_row))]
                            data_rows = full_data
                        data = []
                        for row in data_rows:
                            row_dict = {}
                            for idx, val in enumerate(row):
                                if idx < len(headers):
                                    row_dict[headers[idx]] = val
                            data.append(row_dict)
                        sheets.append({
                            'name': sheet_name,
                            'data': data,
                            'headers': headers,
                            'fullData': full_data,
                            'merged_ranges': merged_ranges,
                            'total_rows': len(full_data),
                            'total_columns': max_col
                        })
                    else:
                        sheets.append({
                            'name': sheet_name,
                            'data': [],
                            'headers': [],
                            'fullData': [],
                            'merged_ranges': [],
                            'total_rows': 0,
                            'total_columns': 0
                        })
                wb.close()
                return {'success': True, 'sheets': sheets}
            else:
                # If it's not a file path, treat it as bytes
                return parse_full_workbook_from_bytes(file_path_or_bytes, instrument_type, max_rows)
        else:
            return parse_full_workbook_from_bytes(file_path_or_bytes, instrument_type, max_rows)
    except Exception as e:
        return {'success': False, 'error': str(e), 'sheets': []}


def parse_full_workbook_from_bytes(file_bytes, instrument_type='money-market', max_rows=100000):
    """
    Parse an Excel file from bytes (e.g., uploaded file) with the same logic.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            full_data = []
            max_col = 0
            for row in ws.iter_rows(values_only=True, max_row=max_rows):
                if not any(cell is not None for cell in row):
                    continue
                row_list = [cell if cell is not None else '' for cell in row]
                full_data.append(row_list)
                if len(row_list) > max_col:
                    max_col = len(row_list)
            # Pad rows
            for row in full_data:
                if len(row) < max_col:
                    row.extend([''] * (max_col - len(row)))

            merged_ranges = []
            for merge in ws.merged_cells.ranges:
                merged_ranges.append({
                    'min_row': merge.min_row - 1,
                    'max_row': merge.max_row - 1,
                    'min_col': merge.min_col - 1,
                    'max_col': merge.max_col - 1
                })

            if full_data:
                first_row = full_data[0]
                is_header_row = any(isinstance(cell, str) and not cell.isdigit() for cell in first_row)
                if is_header_row:
                    headers = [str(h) for h in first_row]
                    data_rows = full_data[1:]
                else:
                    headers = [f"Column_{i+1}" for i in range(len(first_row))]
                    data_rows = full_data
                data = []
                for row in data_rows:
                    row_dict = {}
                    for idx, val in enumerate(row):
                        if idx < len(headers):
                            row_dict[headers[idx]] = val
                    data.append(row_dict)
                sheets.append({
                    'name': sheet_name,
                    'data': data,
                    'headers': headers,
                    'fullData': full_data,
                    'merged_ranges': merged_ranges,
                    'total_rows': len(full_data),
                    'total_columns': max_col
                })
            else:
                sheets.append({
                    'name': sheet_name,
                    'data': [],
                    'headers': [],
                    'fullData': [],
                    'merged_ranges': [],
                    'total_rows': 0,
                    'total_columns': 0
                })
        wb.close()
        return {'success': True, 'sheets': sheets}
    except Exception as e:
        # Fallback to pandas for problematic files (e.g., .xlsb or corrupted)
        try:
            df_dict = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, header=None, dtype=str)
            sheets = []
            for sheet_name, df in df_dict.items():
                full_data = df.values.tolist()
                full_data = [[str(cell) if pd.notna(cell) else '' for cell in row] for row in full_data]
                # No merged ranges from pandas
                merged_ranges = []
                if full_data:
                    first_row = full_data[0]
                    is_header_row = any(isinstance(cell, str) and not cell.isdigit() for cell in first_row)
                    if is_header_row:
                        headers = [str(h) for h in first_row]
                        data_rows = full_data[1:]
                    else:
                        headers = [f"Column_{i+1}" for i in range(len(first_row))]
                        data_rows = full_data
                    data = []
                    for row in data_rows:
                        row_dict = {}
                        for idx, val in enumerate(row):
                            if idx < len(headers):
                                row_dict[headers[idx]] = val
                        data.append(row_dict)
                    sheets.append({
                        'name': sheet_name,
                        'data': data,
                        'headers': headers,
                        'fullData': full_data,
                        'merged_ranges': merged_ranges,
                        'total_rows': len(full_data),
                        'total_columns': len(full_data[0]) if full_data else 0
                    })
                else:
                    sheets.append({
                        'name': sheet_name,
                        'data': [],
                        'headers': [],
                        'fullData': [],
                        'merged_ranges': [],
                        'total_rows': 0,
                        'total_columns': 0
                    })
            return {'success': True, 'sheets': sheets}
        except Exception as e2:
            return {'success': False, 'error': str(e2), 'sheets': []}