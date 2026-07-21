import json
from flask import request, jsonify
from utils.db import get_db
import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
import tempfile
import os
import sys
# Use our central excel_parser
from utils.excel_parser import parse_full_workbook

def create_processed_data_table():
    conn = get_db()
    if not conn:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS processed_data (
                id INT AUTO_INCREMENT PRIMARY KEY,
                dataset_id INT NOT NULL,
                instrument_type VARCHAR(50) NOT NULL,
                original_data JSON,
                cleaned_data JSON,
                cleaning_options JSON,
                processing_stats JSON,
                processed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_dataset_id (dataset_id)
            )
        """)
        conn.commit()
        cursor.close()
        conn.close()
        return True
    except Exception as e:
        print(f"Error creating processed_data table: {e}")
        conn.close()
        return False

def extract_table_from_sheet(sheet):
    data = []
    start_row = None
    start_col = None
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, col)
            if cell.value is not None:
                start_row = row
                start_col = col
                break
        if start_row is not None:
            break
    if start_row is None:
        return []
    headers = []
    for col in range(start_col, sheet.max_column + 1):
        cell = sheet.cell(start_row, col)
        if cell.value is not None and str(cell.value).strip():
            headers.append(str(cell.value).strip())
        else:
            break
    if not headers:
        return []
    for row in range(start_row + 1, sheet.max_row + 1):
        row_data = {}
        for i, col in enumerate(range(start_col, start_col + len(headers))):
            cell = sheet.cell(row, col)
            row_data[headers[i]] = cell.value
        if any(v is not None for v in row_data.values()):
            data.append(row_data)
    return data

def extract_labelled_values(sheet):
    pairs = {}
    keywords = ['rate', 'face value', 'coupon', 'maturity', 'discount', 'interest', 'yield', 'amount', 'principal']
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, col)
            if cell.value and isinstance(cell.value, str):
                label = cell.value.strip().lower()
                if any(k in label for k in keywords):
                    for dr, dc in [(0,1), (1,0), (0,-1), (-1,0)]:
                        try:
                            val_cell = sheet.cell(row + dr, col + dc)
                            if val_cell.value is not None and isinstance(val_cell.value, (int, float)):
                                pairs[label] = val_cell.value
                                break
                        except:
                            pass
    return pairs

def parse_intelligent_excel(file_path, instrument_type):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    all_rows = []
    required_mapping = {
        'money-market': {
            'Rate': ['rate', 'interest rate', 'yield'],
            'Amount': ['amount', 'principal', 'face value'],
            'Instrument': ['instrument', 'name', 'security']
        },
        'bonds': {
            'CouponRate': ['coupon rate', 'coupon', 'rate'],
            'FaceValue': ['face value', 'face', 'principal'],
            'Yield': ['yield', 'yield to maturity', 'ytm'],
            'BondName': ['bond name', 'name', 'security']
        },
        'tbills': {
            'DiscountRate': ['discount rate', 'rate', 'discount'],
            'FaceValue': ['face value', 'face', 'principal'],
            'TBillName': ['tbill name', 'name', 'security']
        }
    }
    mapping = required_mapping.get(instrument_type, {})
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        table = extract_table_from_sheet(sheet)
        if table:
            all_rows.extend(table)
        labels = extract_labelled_values(sheet)
        if labels:
            if not table:
                all_rows.append(labels)
            else:
                for row in all_rows:
                    for k, v in labels.items():
                        if k not in row:
                            row[k] = v
    mapped_rows = []
    for row in all_rows:
        new_row = {}
        for req_col, search_terms in mapping.items():
            found = None
            if req_col in row:
                found = row[req_col]
            else:
                for col, val in row.items():
                    if col and isinstance(col, str):
                        if any(term in col.lower() for term in search_terms):
                            found = val
                            break
            new_row[req_col] = found if found is not None else ''
        mapped_rows.append(new_row)
    
    workbook.close()
    
    return {
        'sheets': [{
            'name': workbook.sheetnames[0] if workbook.sheetnames else 'Sheet1',
            'headers': list(mapping.keys()) if mapping else [],
            'data': mapped_rows,
            'row_count': len(mapped_rows),
            'column_count': len(mapping.keys()) if mapping else 0
        }],
        'metadata': {
            'sheet_names': workbook.sheetnames,
            'sheet_count': len(workbook.sheetnames),
            'total_rows': len(mapped_rows),
            'total_columns': len(mapping.keys()) if mapping else 0
        },
        'warnings': []
    }

def clean_data(data, cleaning_options):
    if not data or not isinstance(data, list):
        return {'cleaned_data': [], 'stats': {}}
    df = pd.DataFrame(data)
    original_rows = len(df)
    stats = {
        'original_rows': original_rows,
        'removed_duplicates': 0,
        'filled_missing_text': 0,
        'removed_missing_rows': 0,
        'trimmed_whitespace': 0,
        'final_rows': original_rows
    }
    if cleaning_options.get('removeDuplicates', True):
        before = len(df)
        df = df.drop_duplicates()
        stats['removed_duplicates'] = before - len(df)
    if cleaning_options.get('fillMissingText', True):
        text_cols = df.select_dtypes(include=['object']).columns
        for col in text_cols:
            filled = df[col].fillna('').count()
            stats['filled_missing_text'] += filled
    if cleaning_options.get('dropRowsWithMissing', False):
        before = len(df)
        df = df.dropna()
        stats['removed_missing_rows'] = before - len(df)
    if cleaning_options.get('trimWhitespace', True):
        text_cols = df.select_dtypes(include=['object']).columns
        for col in text_cols:
            df[col] = df[col].astype(str).str.strip()
            stats['trimmed_whitespace'] = len(df)
    cleaned_data = df.to_dict('records')
    stats['final_rows'] = len(cleaned_data)
    return {'cleaned_data': cleaned_data, 'stats': stats}

def validate_row(row, required_fields):
    missing = []
    invalid = []
    for field in required_fields:
        value = row.get(field)
        if value is None or value == '' or value == 0:
            missing.append(field)
    numeric_fields = ['Principal', 'FaceValue', 'InterestRate', 'CouponRate', 'DiscountRate', 'DaysToMaturity', 'Yield', 'Price']
    for field in numeric_fields:
        if field in row:
            try:
                float(row[field])
            except (ValueError, TypeError):
                invalid.append({'field': field, 'error': 'Must be a number'})
    return {'valid': len(missing) == 0 and len(invalid) == 0, 'missing': missing, 'invalid': invalid}

def apply_mapping(data, column_mapping):
    if not data or not column_mapping:
        return data
    mapped = []
    for row in data:
        new = {}
        for sys_col, file_col in column_mapping.items():
            new[sys_col] = row.get(file_col, '')
        mapped.append(new)
    return mapped

def auto_match_columns(file_columns, required_columns):
    keywords = {
        'Principal': ['principal', 'principal amount', 'investment', 'amount', 'notional'],
        'InterestRate': ['interest rate', 'rate', 'interest', 'coupon'],
        'InvestmentAmount': ['investment amount', 'investment', 'amount'],
        'DaysToMaturity': ['days to maturity', 'maturity days', 'term', 'tenor'],
        'CouponRate': ['coupon rate', 'coupon', 'rate'],
        'CouponFrequency': ['coupon frequency', 'frequency', 'payment frequency'],
        'FaceValue': ['face value', 'face', 'par value', 'par', 'amount'],
        'Yield': ['yield', 'ytm', 'yield to maturity'],
        'SettlementDate': ['settlement date', 'settlement', 'trade date'],
        'MaturityDate': ['maturity date', 'maturity', 'due date'],
        'DiscountRate': ['discount rate', 'discount'],
        'PurchasePrice': ['purchase price', 'purchase', 'price'],
        'RedemptionValue': ['redemption value', 'redemption']
    }
    mapping = {}
    for sys_col in required_columns:
        for file_col in file_columns:
            if any(kw in str(file_col).lower() for kw in keywords.get(sys_col, [])):
                mapping[sys_col] = file_col
                break
    return mapping

# Cache for parsed workbook data
_workbook_cache = {}

def data_processing_routes(app):
    create_processed_data_table()

    @app.route('/api/data/workbook/<file_id>', methods=['GET'])
    def get_workbook_by_id(file_id):
        print(f"=== Get Workbook by ID: {file_id} ===")
        
        if file_id in _workbook_cache:
            print(f"✅ Returning cached workbook data for {file_id}")
            return jsonify({'success': True, 'data': _workbook_cache[file_id]})
        
        upload_folder = 'uploads'
        if not os.path.exists(upload_folder):
            return jsonify({'success': False, 'error': 'Uploads folder not found'}), 404
        
        matching_files = []
        for filename in os.listdir(upload_folder):
            if filename.startswith(file_id):
                file_path = os.path.join(upload_folder, filename)
                matching_files.append(file_path)
        
        if not matching_files:
            print(f"ERROR: No file found with ID {file_id}")
            return jsonify({'success': False, 'error': 'File not found'}), 404
        
        file_path = matching_files[0]
        print(f"Found file: {file_path}")
        
        instrument_type = request.args.get('instrument_type', 'money-market')
        
        try:
            # Use our central parser
            result = parse_full_workbook(file_path, instrument_type, max_rows=10000)
            print(f"Workbook parsed: {len(result.get('sheets', []))} sheets")
            
            _workbook_cache[file_id] = result
            print(f"✅ Cached workbook data for {file_id}")
            
            return jsonify({'success': True, 'data': result})
        except Exception as e:
            print(f"ERROR parsing workbook: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/api/data/parse-excel', methods=['POST', 'OPTIONS'])
    def parse_excel_endpoint():
        if request.method == 'OPTIONS':
            response = jsonify({'success': True})
            response.headers.add('Access-Control-Allow-Origin', '*')
            response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
            response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
            return response, 200
        
        print("=== Parse Excel Request ===")
        print(f"Request method: {request.method}")
        print(f"Request files keys: {list(request.files.keys())}")
        print(f"Request form keys: {list(request.form.keys())}")
        
        file = request.files.get('file')
        instrument_type = request.form.get('instrument_type', 'money-market')
        return_full_workbook = request.form.get('return_full_workbook', 'false').lower() == 'true'
        
        print(f"File received: {file.filename if file else 'None'}")
        print(f"File content type: {file.content_type if file else 'None'}")
        print(f"File size: {len(file.read()) if file else 0} bytes")
        if file:
            file.seek(0)
        print(f"Instrument type: {instrument_type}")
        print(f"Return full workbook: {return_full_workbook}")
        
        if not file:
            print("ERROR: No file uploaded")
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
        
        import uuid
        file_id = str(uuid.uuid4())
        upload_timestamp = datetime.now().isoformat()
        
        upload_folder = 'uploads'
        os.makedirs(upload_folder, exist_ok=True)
        file_path = os.path.join(upload_folder, f"{file_id}_{file.filename}")
        file.save(file_path)
        print(f"File saved to: {file_path} with ID: {file_id}")
        
        try:
            if return_full_workbook:
                print("Calling parse_full_workbook (for viewer)...")
                result = parse_full_workbook(file_path, instrument_type, max_rows=10000)
                print(f"Parse result type: {type(result)}")
                print(f"Parse result sheets: {len(result.get('sheets', []))}")
                _workbook_cache[file_id] = result
                print(f"✅ Cached workbook data for {file_id}")
            else:
                print("Calling parse_intelligent_excel (fast)...")
                result = parse_intelligent_excel(file_path, instrument_type)
                print(f"Parse result type: {type(result)}")
                print(f"Parse result sheets: {len(result.get('sheets', []))}")
                _workbook_cache[file_id] = result
                print(f"✅ Cached workbook data for {file_id}")
            
            if result and isinstance(result, dict):
                result['metadata'] = result.get('metadata', {})
                result['metadata']['file_id'] = file_id
                result['metadata']['original_filename'] = file.filename
                result['metadata']['upload_timestamp'] = upload_timestamp
                result['metadata']['file_path'] = file_path
            else:
                print(f"ERROR: parsing returned invalid type: {type(result)}")
                result = {'sheets': [], 'metadata': {'file_id': file_id, 'original_filename': file.filename, 'upload_timestamp': upload_timestamp}, 'warnings': ['Parsing failed']}
            
            if not result or not isinstance(result, dict):
                print("ERROR: Invalid parsing result")
                result = {'sheets': [], 'metadata': {'file_id': file_id, 'original_filename': file.filename, 'upload_timestamp': upload_timestamp}, 'warnings': ['Invalid parsing result']}
            
            if 'sheets' not in result:
                result['sheets'] = []
            if 'metadata' not in result:
                result['metadata'] = {}
            if 'warnings' not in result:
                result['warnings'] = []
            
            print(f"Final result: {len(result.get('sheets', []))} sheets, {len(result.get('warnings', []))} warnings")
                
        except Exception as e:
            print(f"ERROR during parsing: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': f'Parsing error: {str(e)}'}), 500
        
        response = jsonify({'success': True, 'data': result})
        response.headers.add('Access-Control-Allow-Origin', '*')
        print("=== Parse Excel Complete ===")
        return response

    @app.route('/api/data/clean', methods=['POST', 'OPTIONS'])
    def clean_dataset():
        if request.method == 'OPTIONS':
            return '', 200
        payload = request.get_json() or {}
        data = payload.get('data', [])
        cleaning_options = payload.get('cleaning_options', {})
        result = clean_data(data, cleaning_options)
        return jsonify({'success': True, 'data': result})

    @app.route('/api/data/validate', methods=['POST', 'OPTIONS'])
    def validate_dataset():
        if request.method == 'OPTIONS':
            return '', 200
        payload = request.get_json() or {}
        data = payload.get('data', [])
        required_fields = payload.get('required_fields', [])
        results = []
        for idx, row in enumerate(data):
            val = validate_row(row, required_fields)
            results.append({'row_index': idx, **val})
        valid = sum(1 for r in results if r['valid'])
        return jsonify({
            'success': True,
            'data': {
                'total_rows': len(results),
                'valid_rows': valid,
                'invalid_rows': len(results) - valid,
                'results': results
            }
        })

    @app.route('/api/data/apply-mapping', methods=['POST', 'OPTIONS'])
    def apply_column_mapping():
        if request.method == 'OPTIONS':
            return '', 200
        payload = request.get_json() or {}
        data = payload.get('data', [])
        column_mapping = payload.get('column_mapping', {})
        mapped = apply_mapping(data, column_mapping)
        return jsonify({'success': True, 'data': mapped})

    @app.route('/api/data/auto-match', methods=['POST', 'OPTIONS'])
    def auto_match_columns_endpoint():
        if request.method == 'OPTIONS':
            return '', 200
        payload = request.get_json() or {}
        file_columns = payload.get('file_columns', [])
        required_columns = payload.get('required_columns', [])
        mapping = auto_match_columns(file_columns, required_columns)
        return jsonify({'success': True, 'data': mapping})

    @app.route('/api/data/process', methods=['POST', 'OPTIONS'])
    def process_dataset():
        if request.method == 'OPTIONS':
            return '', 200
        payload = request.get_json() or {}
        dataset_id = payload.get('dataset_id')
        data = payload.get('data', [])
        column_mapping = payload.get('column_mapping', {})
        cleaning_options = payload.get('cleaning_options', {})
        required_fields = payload.get('required_fields', [])
        instrument_type = payload.get('instrument_type', 'money-market')

        mapped = apply_mapping(data, column_mapping)
        clean_result = clean_data(mapped, cleaning_options)
        cleaned = clean_result['cleaned_data']
        validations = []
        for idx, row in enumerate(cleaned):
            val = validate_row(row, required_fields)
            validations.append({'row_index': idx, **val})
        valid_count = sum(1 for v in validations if v['valid'])

        conn = get_db()
        processed_id = None
        if conn:
            try:
                cursor = conn.cursor()
                cursor.execute(
                    """INSERT INTO processed_data 
                       (dataset_id, instrument_type, original_data, cleaned_data, cleaning_options, processing_stats) 
                       VALUES (%s, %s, %s, %s, %s, %s)""",
                    (dataset_id, instrument_type, json.dumps(data), json.dumps(cleaned),
                     json.dumps(cleaning_options), json.dumps(clean_result['stats']))
                )
                conn.commit()
                processed_id = cursor.lastrowid
                cursor.close()
                conn.close()
            except Exception as e:
                print(f"Error saving processed data: {e}")
                conn.close()

        return jsonify({
            'success': True,
            'data': {
                'processed_id': processed_id,
                'cleaned_data': cleaned,
                'stats': clean_result['stats'],
                'validation': {
                    'total_rows': len(validations),
                    'valid_rows': valid_count,
                    'invalid_rows': len(validations) - valid_count,
                    'results': validations
                }
            }
        })