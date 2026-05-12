import requests
import openpyxl
from io import BytesIO
import os

def create_test_excel(rows, headers, data_rows):
    """Create Excel file with dynamic data from environment"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Add headers from environment
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    
    # Add data rows from environment
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, '')
            ws.cell(row_idx, col_idx, value)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def test_upload():
    api_url = os.environ.get('API_URL', 'http://localhost:5000/api/upload')
    
    # Get test data from environment
    headers_json = os.environ.get('EXCEL_HEADERS', '')
    data_json = os.environ.get('TEST_DATA', '')
    
    if not headers_json or not data_json:
        print("ERROR: EXCEL_HEADERS and TEST_DATA required")
        return
    
    headers = json.loads(headers_json)
    test_data = json.loads(data_json)
    rows = len(test_data)
    
    print(f"Testing upload with {rows} rows")
    print("=" * 30)
    
    buffer = create_test_excel(rows, headers, test_data)
    
    try:
        response = requests.post(
            api_url,
            files={'file': ('test.xlsx', buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
            data={'instrument_type': 'treasury_bills'}
        )
        
        print(f"Status: {response.status_code}")
        
        if response.status_code == 200:
            data = response.json().get('data', {})
            rows_returned = len(data.get('data', []))
            print(f"Success: True")
            print(f"Rows: {rows_returned}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    import json
    test_upload()