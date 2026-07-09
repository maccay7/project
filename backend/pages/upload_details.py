import os
import json
import base64
import uuid
import io
from openpyxl import load_workbook
from utils.db import get_db


def parse_upload_file(file_storage):
    filename = file_storage.filename.lower()
    binary = file_storage.read()
    
    try:
        if filename.endswith('.csv'):
            text = binary.decode('utf-8')
            lines = text.split('\n')
            if not lines or not lines[0].strip():
                return {
                    'success': False,
                    'error': 'Empty file',
                    'file_base64': base64.b64encode(binary).decode('utf-8')
                }
            
            first_line = lines[0]
            delimiter = ','
            if ';' in first_line and ',' not in first_line:
                delimiter = ';'
            elif '\t' in first_line:
                delimiter = '\t'
            
            headers = [h.strip().replace('"', '') for h in first_line.split(delimiter)]
            data = []
            for line in lines[1:]:
                if not line.strip():
                    continue
                values = [v.strip().replace('"', '') for v in line.split(delimiter)]
                row = {}
                for i, header in enumerate(headers):
                    row[header] = values[i] if i < len(values) else ''
                data.append(row)
                
        elif filename.endswith(('.xlsx', '.xls', '.xlsm', '.xlsb')):
            wb = load_workbook(io.BytesIO(binary), data_only=True)
            ws = wb.active
            
            data = []
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(cell) if cell is not None else '' for cell in row]
                else:
                    row_dict = {}
                    for i, cell in enumerate(row):
                        header = headers[i] if i < len(headers) else f'Column_{i}'
                        row_dict[header] = cell if cell is not None else ''
                    data.append(row_dict)
        else:
            return {
                'success': False,
                'error': 'Unsupported file format',
                'file_base64': base64.b64encode(binary).decode('utf-8')
            }
        
        return {
            'success': True,
            'data': data,
            'headers': headers or [],
            'rows': len(data),
            'file_base64': base64.b64encode(binary).decode('utf-8')
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'file_base64': base64.b64encode(binary).decode('utf-8')
        }


def clean_data(data, options):
    if not data or not isinstance(data, list):
        return [], {'total_rows': 0, 'valid_rows': 0, 'removed_rows': 0, 'fixed_missing': 0}
    
    original_rows = len(data)
    stats = {
        'total_rows': original_rows,
        'removed_duplicates': 0,
        'filled_missing_text': 0,
        'removed_missing_rows': 0,
        'trimmed_whitespace': 0,
        'final_rows': original_rows
    }
    
    if options.get('removeDuplicates', True):
        seen = set()
        unique_data = []
        for row in data:
            row_key = json.dumps(row, sort_keys=True)
            if row_key not in seen:
                seen.add(row_key)
                unique_data.append(row)
        stats['removed_duplicates'] = original_rows - len(unique_data)
        data = unique_data
    
    if options.get('trimWhitespace', True):
        for row in data:
            for key in row:
                if isinstance(row[key], str):
                    row[key] = row[key].strip()
        stats['trimmed_whitespace'] = len(data)
    
    if options.get('convertToNumbers', True):
        for row in data:
            for key in row:
                if isinstance(row[key], str):
                    try:
                        row[key] = float(row[key])
                    except (ValueError, TypeError):
                        pass
    
    if options.get('fillMissingText', True):
        fill_value = options.get('fill_missing', '')
        for row in data:
            for key in row:
                if row[key] is None or row[key] == '':
                    row[key] = fill_value
                    stats['filled_missing_text'] += 1
    
    if options.get('dropRowsWithMissing', False):
        before_drop = len(data)
        data = [row for row in data if all(v is not None and v != '' for v in row.values())]
        stats['removed_missing_rows'] = before_drop - len(data)
    
    if options.get('removeOutliers', False):
        numeric_cols = set()
        for row in data:
            for key, val in row.items():
                if isinstance(val, (int, float)):
                    numeric_cols.add(key)
        
        for col in numeric_cols:
            values = [row[col] for row in data if isinstance(row[col], (int, float))]
            if len(values) < 2:
                continue
            mean = sum(values) / len(values)
            variance = sum((v - mean) ** 2 for v in values) / len(values)
            std = variance ** 0.5
            if std > 0:
                threshold = 3 * std
                data = [row for row in data if abs(row[col] - mean) <= threshold if isinstance(row[col], (int, float))]
    
    if options.get('standardizeDates', False):
        for row in data:
            for key in row:
                if 'date' in key.lower() and row[key]:
                    try:
                        from datetime import datetime
                        d = datetime.strptime(str(row[key]), '%Y-%m-%d')
                        row[key] = d.strftime('%Y-%m-%d')
                    except:
                        pass
    
    if options.get('removeSpecialChars', False):
        import re
        for row in data:
            for key in row:
                if isinstance(row[key], str):
                    row[key] = re.sub(r'[^a-zA-Z0-9\s]', '', row[key])
    
    if options.get('changeCase', False):
        case_type = options.get('caseType', 'none')
        for row in data:
            for key in row:
                if isinstance(row[key], str):
                    if case_type == 'upper':
                        row[key] = row[key].upper()
                    elif case_type == 'lower':
                        row[key] = row[key].lower()
                    elif case_type == 'title':
                        row[key] = row[key].title()
    
    if options.get('fillWithCustom', False) and options.get('customFillValue'):
        custom_val = options.get('customFillValue')
        for row in data:
            for key in row:
                if row[key] is None or row[key] == '':
                    row[key] = custom_val
    
    if options.get('removeColumnsAllMissing', False):
        cols_to_keep = []
        if data:
            for col in data[0].keys():
                if any(row.get(col) is not None and row.get(col) != '' for row in data):
                    cols_to_keep.append(col)
        data = [{k: row[k] for k in cols_to_keep} for row in data]
    
    if options.get('capOutliers', False):
        numeric_cols = set()
        for row in data:
            for key, val in row.items():
                if isinstance(val, (int, float)):
                    numeric_cols.add(key)
        
        for col in numeric_cols:
            values = [row[col] for row in data if isinstance(row[col], (int, float))]
            if len(values) < 2:
                continue
            mean = sum(values) / len(values)
            variance = sum((v - mean) ** 2 for v in values) / len(values)
            std = variance ** 0.5
            if std > 0:
                upper = mean + 3 * std
                lower = mean - 3 * std
                for row in data:
                    if isinstance(row[col], (int, float)):
                        row[col] = max(lower, min(upper, row[col]))
    
    if options.get('removeRowsSpecificColumnEmpty', False) and options.get('specificColumn'):
        col = options.get('specificColumn')
        data = [row for row in data if row.get(col) is not None and row.get(col) != '']
    
    if options.get('standardizeNumericRange', False):
        numeric_cols = set()
        for row in data:
            for key, val in row.items():
                if isinstance(val, (int, float)):
                    numeric_cols.add(key)
        
        for col in numeric_cols:
            values = [row[col] for row in data if isinstance(row[col], (int, float))]
            if len(values) < 2:
                continue
            min_val = min(values)
            max_val = max(values)
            if max_val != min_val:
                for row in data:
                    if isinstance(row[col], (int, float)):
                        row[col] = (row[col] - min_val) / (max_val - min_val)
    
    if options.get('removeEmptyRows', False):
        data = [row for row in data if any(v is not None and v != '' for v in row.values())]
    
    if options.get('fillForward', False):
        if data:
            for i in range(1, len(data)):
                for key in data[i]:
                    if data[i][key] is None or data[i][key] == '':
                        data[i][key] = data[i-1].get(key, '')
    
    if options.get('fillBackward', False):
        if data:
            for i in range(len(data) - 2, -1, -1):
                for key in data[i]:
                    if data[i][key] is None or data[i][key] == '':
                        data[i][key] = data[i+1].get(key, '')
    
    stats['final_rows'] = len(data)
    stats['valid_rows'] = len(data)
    stats['removed_rows'] = original_rows - len(data)
    
    return data, stats


def save_dataset(name, file_base64='', sheet_names=None, upload_id=None, data=None, headers=None, instrument_type=None):
    conn = get_db()
    if not conn:
        return None
    try:
        ds_id = upload_id or str(uuid.uuid4())
        cursor = conn.cursor()
        cursor.execute(
            "REPLACE INTO datasets (id, name, file_base64, data, headers, instrument_type, upload_status, done) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
            (
                ds_id,
                name,
                file_base64,
                json.dumps(data) if data is not None else None,
                json.dumps(headers) if headers is not None else None,
                instrument_type,
                'uploaded',
                False
            )
        )
        conn.commit()
        cursor.close()
        conn.close()
        return {'id': ds_id, 'name': name, 'file_base64': file_base64, 'data': data, 'headers': headers, 'instrument_type': instrument_type}
    except Exception as e:
        print(f"Save dataset error: {e}")
        return None


def get_saved_datasets():
    conn = get_db()
    if not conn:
        return []
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id, name, JSON_LENGTH(data) as rows, instrument_type, done, created_at FROM datasets WHERE done = FALSE ORDER BY created_at DESC")
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return [
            {
                'id': r.get('id'),
                'name': r.get('name'),
                'rows': r.get('rows') or 0,
                'instrument_type': r.get('instrument_type'),
                'done': bool(r.get('done'))
            }
            for r in rows
        ]
    except Exception as e:
        print(f"Get datasets error: {e}")
        return []


def load_saved_dataset(dataset_id):
    if not dataset_id:
        return None
    conn = get_db()
    if not conn:
        return None
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM datasets WHERE id = %s", (dataset_id,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        if not row:
            return None
        data = row.get('data')
        try:
            data = json.loads(data) if isinstance(data, str) else data
        except Exception:
            data = None
        headers = row.get('headers')
        try:
            headers = json.loads(headers) if isinstance(headers, str) else headers
        except Exception:
            headers = None
        return {
            'id': row.get('id'),
            'name': row.get('name'),
            'file_base64': row.get('file_base64'),
            'data': data,
            'headers': headers,
            'instrument_type': row.get('instrument_type'),
            'done': bool(row.get('done'))
        }
    except Exception as e:
        print(f"Load dataset error: {e}")
        return None


def delete_saved_dataset(dataset_id):
    if not dataset_id:
        return False
    conn = get_db()
    if not conn:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM datasets WHERE id = %s", (dataset_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return True
    except Exception as e:
        print(f"Delete dataset error: {e}")
        return False


def mark_dataset_done(dataset_id, done=True):
    if not dataset_id:
        return False
    conn = get_db()
    if not conn:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute("UPDATE datasets SET done = %s WHERE id = %s", (1 if done else 0, dataset_id))
        conn.commit()
        cursor.close()
        conn.close()
        return True
    except Exception as e:
        print(f"Mark done error: {e}")
        return False