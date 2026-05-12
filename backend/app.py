from flask import Flask, request, jsonify
from flask_cors import CORS
import pymysql
import requests
import json
import openpyxl
import os
import base64
import tempfile
import uuid
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()
app = Flask(__name__)
CORS(app, origins='*', supports_credentials=True)

# ==================== CONFIGURATION ====================
DB_CONFIG = {
    'host': os.environ.get('DB_HOST', 'localhost'),
    'user': os.environ.get('DB_USER', 'root'),
    'password': os.environ.get('DB_PASSWORD', ''),
    'database': os.environ.get('DB_NAME', 'duracapital'),
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor
}

FRED_API_KEY = os.environ.get('FRED_API_KEY', '')
FRED_BASE_URL = 'https://api.stlouisfed.org/fred'
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', '')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', '')

# ==================== DATABASE FUNCTIONS ====================

def get_db():
    try:
        return pymysql.connect(**DB_CONFIG)
    except Exception as e:
        print(f"DB error: {e}")
        return None

# ==================== FRED YIELD CURVE ====================

def fetch_fred_data(series_id):
    if not FRED_API_KEY:
        return None
    try:
        url = f"{FRED_BASE_URL}/series/observations?series_id={series_id}&api_key={FRED_API_KEY}&file_type=json&limit=10"
        resp = requests.get(url, timeout=10)
        return resp.json() if resp.status_code == 200 else None
    except:
        return None

def get_yield_curve_rate(maturity_years):
    """Get yield curve rate for given maturity from FRED API"""
    if not FRED_API_KEY:
        return None
    
    series_map = {
        0.25: 'TB3MS',   # 3-month
        0.5: 'TB6MS',    # 6-month
        1: 'DGS1',       # 1-year
        2: 'DGS2',       # 2-year
        5: 'DGS5',       # 5-year
        10: 'DGS10',     # 10-year
        30: 'DGS30'      # 30-year
    }
    
    closest = min(series_map.keys(), key=lambda x: abs(x - maturity_years))
    series_id = series_map[closest]
    
    data = fetch_fred_data(series_id)
    if data and data.get('observations'):
        return float(data['observations'][0]['value'])
    return None

# ==================== CALCULATIONS ====================

def calculate_treasury_bill(row):
    """Treasury Bill calculations with yield curve integration"""
    face_value = float(row.get('faceValue', row.get('face_value', 0)))
    purchase_price = float(row.get('purchasePrice', row.get('purchase_price', 0)))
    days = float(row.get('daysToMaturity', row.get('days_to_maturity', 0)))
    
    if face_value <= 0 or purchase_price <= 0 or days <= 0:
        return {**row, 'error': 'Missing required fields'}
    
    discount = face_value - purchase_price
    years = days / 365
    yield_curve_rate = get_yield_curve_rate(years)
    
    # Calculations
    discount_yield = (discount / face_value) * (360 / days) * 100
    bond_yield = (discount / purchase_price) * (365 / days) * 100
    money_market_yield = (discount / purchase_price) * (360 / days) * 100
    price_per_100 = (purchase_price / face_value) * 100
    
    return {
        **row,
        'instrument_type': 'Treasury Bill',
        'face_value': face_value,
        'purchase_price': purchase_price,
        'days_to_maturity': days,
        'discount': round(discount, 2),
        'discount_yield': round(discount_yield, 4),
        'bond_equivalent_yield': round(bond_yield, 4),
        'money_market_yield': round(money_market_yield, 4),
        'price_per_100': round(price_per_100, 4),
        'yield_curve_rate': round(yield_curve_rate, 4) if yield_curve_rate else None,
        'yield_spread': round(bond_yield - yield_curve_rate, 4) if yield_curve_rate else None
    }

def calculate_bond(row):
    """Bond calculations with yield curve integration"""
    face_value = float(row.get('faceValue', row.get('face_value', 0)))
    coupon_rate = float(row.get('couponRate', row.get('coupon_rate', 0)))
    current_price = float(row.get('currentPrice', row.get('current_price', 0)))
    years = float(row.get('yearsToMaturity', row.get('years_to_maturity', 0)))
    
    if face_value <= 0 or current_price <= 0 or years <= 0:
        return {**row, 'error': 'Missing required fields'}
    
    coupon_pct = coupon_rate / 100 if coupon_rate > 1 else coupon_rate
    annual_coupon = face_value * coupon_pct
    
    # Current Yield
    current_yield = (annual_coupon / current_price) * 100
    
    # Approximate Yield to Maturity
    ytm = (annual_coupon + (face_value - current_price) / years) / ((face_value + current_price) / 2) * 100
    
    # Yield curve for comparison
    yield_curve_rate = get_yield_curve_rate(years)
    
    # Approximate Duration
    if ytm > 0:
        duration = (1 + ytm/100) / (ytm/100) - (1 + ytm/100 + years * (coupon_pct - ytm/100)) / (coupon_pct * ((1 + ytm/100)**years - 1) + ytm/100)
    else:
        duration = years
    
    return {
        **row,
        'instrument_type': 'Bond',
        'face_value': face_value,
        'coupon_rate': round(coupon_pct * 100, 2),
        'current_price': current_price,
        'annual_coupon': round(annual_coupon, 2),
        'current_yield': round(current_yield, 4),
        'yield_to_maturity': round(ytm, 4),
        'duration': round(duration, 2),
        'years_to_maturity': years,
        'yield_curve_rate': round(yield_curve_rate, 4) if yield_curve_rate else None,
        'yield_spread': round(ytm - yield_curve_rate, 4) if yield_curve_rate else None
    }

def calculate_money_market(row):
    """Money Market calculations with yield curve integration"""
    principal = float(row.get('principal', row.get('Principal', 0)))
    rate = float(row.get('interest_rate', row.get('interestRate', row.get('rate', 0))))
    days = float(row.get('term_days', row.get('days', row.get('term', 0))))
    
    if principal <= 0 or rate <= 0 or days <= 0:
        return {**row, 'error': 'Missing required fields'}
    
    interest = principal * rate * (days / 365)
    annual_yield = (interest / principal) * (365 / days) * 100
    maturity_value = principal + interest
    
    # Yield curve for comparison
    years = days / 365
    yield_curve_rate = get_yield_curve_rate(years)
    
    return {
        **row,
        'instrument_type': 'Money Market',
        'principal': principal,
        'interest_rate': round(rate * 100, 4),
        'term_days': int(days),
        'interest_earned': round(interest, 2),
        'annual_yield': round(annual_yield, 4),
        'maturity_value': round(maturity_value, 2),
        'yield_curve_rate': round(yield_curve_rate, 4) if yield_curve_rate else None,
        'yield_spread': round(annual_yield - yield_curve_rate, 4) if yield_curve_rate else None
    }

def perform_calculation(instrument_type, data):
    """Route to appropriate calculation function"""
    results = []
    for row in data:
        if instrument_type in ['treasury_bills', 'treasury-bills', 'tbills']:
            results.append(calculate_treasury_bill(row))
        elif instrument_type == 'bonds':
            results.append(calculate_bond(row))
        else:
            results.append(calculate_money_market(row))
    return results

# ==================== API ENDPOINTS ====================

@app.route('/')
def home():
    return jsonify({'message': 'DuraCapital API', 'status': 'running', 'fred_configured': bool(FRED_API_KEY)})

@app.route('/api/login', methods=['POST', 'OPTIONS'])
def login():
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        
        print(f"Login attempt: {email}")
        
        # Hardcoded credentials - FIXED
        if email == 'makanakakanyai@gmail.com' and password == 'Business7mogul':
            return jsonify({
                'success': True,
                'token': str(uuid.uuid4()),
                'user': {
                    'id': 1,
                    'email': email,
                    'full_name': 'Makanaka Kanyai',
                    'role': 'admin'
                }
            })
        
        return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
        
    except Exception as e:
        print(f"Login error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/api/upload', methods=['POST', 'OPTIONS'])
def upload():
    if request.method == 'OPTIONS':
        return '', 200
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    file_data = file.read()
    file_base64 = base64.b64encode(file_data).decode('utf-8')
    file_extension = os.path.splitext(file.filename)[1].lower()
    
    # Parse Excel file
    data = []
    headers = []
    if file_extension in ['.xlsx', '.xls', '.xlsm']:
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp:
                tmp.write(file_data)
                tmp_path = tmp.name
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            sheet = wb.active
            for col in range(1, sheet.max_column + 1):
                val = sheet.cell(1, col).value
                if val:
                    headers.append(str(val))
            for row in range(2, min(sheet.max_row + 1, 102)):
                row_data = {}
                for col, header in enumerate(headers, 1):
                    val = sheet.cell(row, col).value
                    row_data[header] = val if val is not None else ''
                if any(row_data.values()):
                    data.append(row_data)
            wb.close()
            os.unlink(tmp_path)
        except Exception as e:
            print(f"Excel error: {e}")
    
    return jsonify({
        'success': True,
        'data': {
            'file_base64': file_base64,
            'file_name': file.filename,
            'file_type': file_extension,
            'data': data[:100],
            'headers': headers,
            'total_rows': len(data),
            'upload_id': str(uuid.uuid4())
        }
    })

@app.route('/api/calculate', methods=['POST', 'OPTIONS'])
def calculate():
    if request.method == 'OPTIONS':
        return '', 200
    
    data = request.get_json()
    instrument_type = data.get('instrument_type', 'money_market')
    calculation_data = data.get('data', [])
    
    if not calculation_data:
        return jsonify({'error': 'No data provided'}), 400
    
    results = perform_calculation(instrument_type, calculation_data)
    return jsonify({'success': True, 'calculations': results, 'instrument_type': instrument_type})

@app.route('/api/clean', methods=['POST', 'OPTIONS'])
def clean():
    if request.method == 'OPTIONS':
        return '', 200
    
    data = request.get_json()
    original = data.get('data', [])
    options = data.get('options', {})
    
    cleaned = []
    for row in original:
        if options.get('removeEmptyRows', True):
            if not any(v for v in row.values() if v and str(v).strip()):
                continue
        clean_row = {k: v.strip() if isinstance(v, str) else v for k, v in row.items()}
        cleaned.append(clean_row)
    
    return jsonify({
        'success': True,
        'data': cleaned,
        'stats': {'original_rows': len(original), 'cleaned_rows': len(cleaned)}
    })

@app.route('/api/fred-yield-curve', methods=['GET', 'OPTIONS'])
def yield_curve():
    if request.method == 'OPTIONS':
        return '', 200
    
    if not FRED_API_KEY:
        return jsonify({'success': False, 'error': 'FRED_API_KEY not configured'}), 503
    
    maturities = [0.25, 0.5, 1, 2, 5, 10, 30]
    labels = ['3M', '6M', '1Y', '2Y', '5Y', '10Y', '30Y']
    rates = []
    
    for maturity in maturities:
        rate = get_yield_curve_rate(maturity)
        rates.append(rate if rate else None)
    
    return jsonify({
        'success': True,
        'data': {
            'labels': labels,
            'current': rates,
            'datasets': [{
                'label': 'Yield Curve',
                'data': rates,
                'borderColor': '#0B2A44',
                'backgroundColor': 'rgba(11, 42, 68, 0.1)',
                'fill': True,
                'tension': 0.4
            }]
        }
    })

# ==================== DATASET ENDPOINTS ====================

@app.route('/api/save-dataset', methods=['POST', 'OPTIONS'])
def save_dataset():
    if request.method == 'OPTIONS':
        return '', 200
    data = request.get_json()
    conn = get_db()
    if conn:
        cursor = conn.cursor()
        cursor.execute("INSERT INTO datasets (name, file_base64, created_at) VALUES (%s, %s, %s) "
                      "ON DUPLICATE KEY UPDATE file_base64 = %s",
                      (data.get('name'), data.get('file_base64'), datetime.now(), data.get('file_base64')))
        conn.commit()
        cursor.close()
        conn.close()
    return jsonify({'success': True})

@app.route('/api/get-datasets', methods=['GET', 'OPTIONS'])
def get_datasets():
    if request.method == 'OPTIONS':
        return '', 200
    conn = get_db()
    datasets = []
    if conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id, name, created_at FROM datasets ORDER BY created_at DESC")
        for row in cursor.fetchall():
            datasets.append({'id': row[0], 'name': row[1], 'timestamp': row[2].isoformat() if row[2] else None})
        cursor.close()
        conn.close()
    return jsonify({'success': True, 'data': datasets})

@app.route('/api/delete-dataset', methods=['POST', 'OPTIONS'])
def delete_dataset():
    if request.method == 'OPTIONS':
        return '', 200
    conn = get_db()
    if conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM datasets WHERE id = %s", (request.get_json().get('dataset_id'),))
        conn.commit()
        cursor.close()
        conn.close()
    return jsonify({'success': True})

# ==================== DASHBOARD ENDPOINTS ====================

@app.route('/api/dashboard/kpi', methods=['GET', 'OPTIONS'])
def kpi():
    if request.method == 'OPTIONS':
        return '', 200
    return jsonify({'success': True, 'data': {'total_datasets': 0, 'system_health': 'Optimal'}})

@app.route('/api/dashboard/recent-activity', methods=['GET', 'OPTIONS'])
def activity():
    if request.method == 'OPTIONS':
        return '', 200
    return jsonify({'success': True, 'data': [{'id': 1, 'text': 'Welcome', 'time': 'Just now'}]})

@app.route('/api/user/profile', methods=['GET', 'OPTIONS'])
def profile():
    if request.method == 'OPTIONS':
        return '', 200
    return jsonify({'success': True, 'data': {'name': os.environ.get('ADMIN_NAME', ''), 
                   'email': ADMIN_EMAIL, 'role': 'admin'}})

@app.route('/api/system/info', methods=['GET', 'OPTIONS'])
def info():
    if request.method == 'OPTIONS':
        return '', 200
    return jsonify({'success': True, 'data': {'version': '1.0.0', 'status': 'Online', 
                   'fred_configured': bool(FRED_API_KEY)}})

@app.route('/api/dashboard/charts', methods=['GET', 'OPTIONS'])
def charts():
    if request.method == 'OPTIONS':
        return '', 200
    return jsonify({'success': True, 'data': {}})

@app.route('/api/user/preferences', methods=['GET', 'OPTIONS'])
def prefs():
    if request.method == 'OPTIONS':
        return '', 200
    return jsonify({'success': True, 'data': {'language': 'English', 'currency': 'USD'}})

@app.route('/api/user/notifications/settings', methods=['GET', 'OPTIONS'])
def notif():
    if request.method == 'OPTIONS':
        return '', 200
    return jsonify({'success': True, 'data': {'emailNotifications': True}})

@app.route('/api/calculations/history', methods=['GET', 'OPTIONS'])
def calc_history():
    if request.method == 'OPTIONS':
        return '', 200
    return jsonify({'success': True, 'data': []})

@app.route('/api/calculations/execute', methods=['POST', 'OPTIONS'])
def calc_execute():
    if request.method == 'OPTIONS':
        return '', 200
    return jsonify({'success': True, 'data': []})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=True, port=port, host='0.0.0.0')