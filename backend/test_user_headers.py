import requests
import openpyxl
from io import BytesIO
import os

# Get settings from environment
API_URL = os.environ.get('API_URL', 'http://localhost:5000/api/upload')
INSTRUMENT_TYPE = os.environ.get('INSTRUMENT_TYPE', 'treasury_bills')
ROWS = int(os.environ.get('TEST_ROWS', '15'))

def create_test_excel():
    """Create a test Excel file with dynamic data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Headers
    headers_env = os.environ.get('EXCEL_HEADERS', '')
    if headers_env:
        headers = headers_env.split(',')
    else:
        headers = ["Days to Maturity", "Years to Maturity", "Rate", "Nominal", "Carrying Value"]
    
    # Add headers
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    
    # Add data rows
    for row in range(2, ROWS + 2):
        for col in range(1, len(headers) + 1):
            header = headers[col-1]
            if "Rate" in header:
                value = f"{4.5 + (row-2) * 0.1}%"
            elif "Value" in header or "Nominal" in header:
                value = f"${1000 + (row-2) * 100}"
            elif "Date" in header:
                value = f"2024-{row-1:02d}-01"
            else:
                value = f"Item_{row-1}_{col}"
            ws.cell(row, col, value)
    
    return wb

def test_upload():
    print("Testing File Upload")
    print("-" * 30)
    print(f"URL: {API_URL}")
    print(f"Rows: {ROWS}")
    
    # Create and upload file
    wb = create_test_excel()
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    files = {'file': ('test_data.xlsx', buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
    data = {'instrument_type': INSTRUMENT_TYPE}
    
    try:
        response = requests.post(API_URL, files=files, data=data)
        print(f"\nStatus: {response.status_code}")
        
        if response.status_code == 200:
            result = response.json()
            print(f"Success: {result.get('success')}")
            
            if result.get('data'):
                file_data = result['data']
                rows = file_data.get('data', [])
                headers = file_data.get('display_headers', [])
                print(f"Rows returned: {len(rows)}")
                print(f"Columns: {len(headers)}")
                
                if rows:
                    print(f"Sample columns: {list(rows[0].keys())[:3]}")
        else:
            print(f"Error: {response.text}")
            
    except Exception as e:
        print(f"Failed: {e}")

if __name__ == "__main__":
    test_upload()