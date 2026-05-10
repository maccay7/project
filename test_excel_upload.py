import requests
import openpyxl
from io import BytesIO
import os
import json

def create_test_excel(test_data):
    """Create Excel file from test data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    if not test_data or len(test_data) == 0:
        print("ERROR: No test data provided")
        return None
    
    headers = list(test_data[0].keys())
    
    # Add headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add data rows
    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, '')
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def test_upload():
    api_url = os.environ.get('API_URL', 'http://localhost:5000/api/upload')
    test_data_json = os.environ.get('TEST_DATA', '')
    
    if not test_data_json:
        print("ERROR: TEST_DATA environment variable not set")
        return
    
    try:
        test_data = json.loads(test_data_json)
    except json.JSONDecodeError:
        print("ERROR: TEST_DATA is not valid JSON")
        return
    
    buffer = create_test_excel(test_data)
    if not buffer:
        return
    
    try:
        response = requests.post(
            api_url,
            files={'file': ('test.xlsx', buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
            data={'instrument_type': 'treasury_bills'}
        )
        print(f"Status: {response.status_code}")
        if response.status_code == 200:
            print("Success: True")
        else:
            print(f"Error: {response.text}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    test_upload()