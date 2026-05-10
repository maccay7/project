import requests
import json
import os

def load_config():
    """Load all configuration from environment variables"""
    return {
        'api_base': os.environ.get('API_BASE_URL', 'http://localhost:5000'),
        'test_data': os.environ.get('TEST_DATA', ''),
        'upload_id': os.environ.get('UPLOAD_ID', '')
    }

def test_upload(config, test_data):
    """Test upload endpoint with actual file upload"""
    print("\n1. Testing Upload Endpoint...")
    
    if not test_data:
        print("  SKIP - No test data available")
        return False
    
    upload_url = f"{config['api_base']}/api/upload"
    
    # Create a simple Excel file from test data
    try:
        import openpyxl
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if test_data:
            headers = list(test_data[0].keys())
            for col, h in enumerate(headers, 1):
                ws.cell(1, col, h)
            
            for row_idx, row in enumerate(test_data, 2):
                for col_idx, h in enumerate(headers, 1):
                    ws.cell(row_idx, col_idx, row.get(h, ''))
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = requests.post(
            upload_url,
            files={'file': ('test.xlsx', buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
            data={'instrument_type': 'treasury_bills'}
        )
        
        if response.status_code == 200:
            result = response.json()
            print(f"  OK - Upload successful")
            print(f"  OK - Rows: {len(result.get('data', {}).get('data', []))}")
            return True
        else:
            print(f"  ERROR - Status: {response.status_code}")
            return False
            
    except Exception as e:
        print(f"  ERROR - {e}")
        return False

def test_clean(config, test_data):
    """Test clean endpoint with actual API call"""
    print("\n2. Testing Cleaning Endpoint...")
    
    if not test_data:
        print("  SKIP - No test data available")
        return False
    
    clean_url = f"{config['api_base']}/api/clean"
    
    cleaning_options = {
        "removeDuplicates": True,
        "fillMissingValues": True,
        "removeEmptyRows": True,
        "standardizeText": True,
        "trimWhitespace": True
    }
    
    payload = {
        "data": test_data,
        "options": cleaning_options
    }
    
    try:
        response = requests.post(clean_url, json=payload)
        
        if response.status_code == 200:
            result = response.json()
            stats = result.get('stats', {})
            print(f"  OK - Cleaning successful")
            print(f"  OK - Original: {stats.get('original_rows')} rows")
            print(f"  OK - Cleaned: {stats.get('cleaned_rows')} rows")
            return True
        else:
            print(f"  ERROR - Status: {response.status_code}")
            return False
            
    except Exception as e:
        print(f"  ERROR - {e}")
        return False

def test_delete(config, upload_id):
    """Test delete endpoint with actual API call"""
    print("\n3. Testing Delete Endpoint...")
    
    if not upload_id:
        print("  SKIP - No upload_id provided")
        return False
    
    delete_url = f"{config['api_base']}/api/delete-dataset"
    payload = {"upload_id": upload_id}
    
    try:
        response = requests.post(delete_url, json=payload)
        
        if response.status_code == 200:
            result = response.json()
            print(f"  OK - {result.get('message', 'Deleted')}")
            return True
        else:
            print(f"  ERROR - Status: {response.status_code}")
            return False
            
    except Exception as e:
        print(f"  ERROR - {e}")
        return False

def main():
    print("Testing Complete Workflow")
    print("=" * 40)
    
    config = load_config()
    
    # Parse test data from environment
    test_data = []
    if config['test_data']:
        try:
            test_data = json.loads(config['test_data'])
        except:
            print("Error: TEST_DATA is not valid JSON")
            return
    
    print(f"\nAPI Base: {config['api_base']}")
    print(f"Test records: {len(test_data)}")
    
    # Run tests
    upload_ok = test_upload(config, test_data)
    
    # Only test clean if we have test data
    clean_ok = test_clean(config, test_data)
    
    # Only test delete if we have an upload_id
    delete_ok = test_delete(config, config['upload_id'])
    
    # Summary
    print("\n" + "=" * 40)
    print("Test Summary:")
    print(f"  Upload: {'OK' if upload_ok else 'FAIL'}")
    print(f"  Clean: {'OK' if clean_ok else 'FAIL'}")
    print(f"  Delete: {'OK' if delete_ok else 'SKIP' if not config['upload_id'] else 'FAIL'}")

if __name__ == "__main__":
    main()