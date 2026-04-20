"""
Parse and clean uploaded valuation files (CSV / Excel first sheet).

Paths must stay under the configured upload directory (caller validates).
"""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any

_PREVIEW_ROWS = 12
_MAX_SCAN_ROWS = 100_000


def _norm_header(h: str) -> str:
    s = (h or "").strip()
    s = re.sub(r"\s+", "_", s)
    return s or "column"


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _is_missing(v: str) -> bool:
    return v is None or str(v).strip() == ""


def _to_float(v: str) -> float | None:
    if _is_missing(v):
        return None
    raw = str(v).replace(",", "").strip()
    try:
        return float(raw)
    except ValueError:
        return None


def _median(values: list[float]) -> float:
    if not values:
        return 0.0
    s = sorted(values)
    mid = len(s) // 2
    if len(s) % 2 == 0:
        return (s[mid - 1] + s[mid]) / 2.0
    return s[mid]


def _read_csv_matrix(path: Path) -> tuple[list[list[str]], str | None]:
    last_err = None
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = path.read_text(encoding=enc)
            break
        except UnicodeDecodeError as e:
            last_err = str(e)
    if text is None:
        return [], last_err or "Could not decode file as text"

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    rows: list[list[str]] = []
    for raw in reader:
        rows.append([_cell_str(c) for c in raw])
    return rows, None


def _read_xlsx_matrix(path: Path) -> tuple[list[list[str]], str | None]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], "Excel support requires openpyxl (pip install openpyxl)"

    try:
        wb = load_workbook(filename=path, read_only=True, data_only=True)
    except Exception as e:
        return [], f"Could not read workbook: {e}"

    try:
        ws = wb[wb.sheetnames[0]]
        rows: list[list[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= _MAX_SCAN_ROWS:
                break
            rows.append([_cell_str(c) for c in row])
        return rows, None
    finally:
        wb.close()


def _load_matrix(path: Path) -> tuple[list[list[str]], str | None, str]:
    suf = path.suffix.lower()
    if suf == ".csv":
        m, err = _read_csv_matrix(path)
        return m, err, "csv"
    if suf in (".xlsx", ".xlsm"):
        m, err = _read_xlsx_matrix(path)
        return m, err, "xlsx"
    if suf == ".xls":
        return [], "Legacy .xls is not supported; save as .xlsx or .csv", "xls"
    if suf == ".xlsb":
        return [], ".xlsb is not supported; export as .xlsx or .csv", "xlsb"
    return [], f"Unsupported extension {suf}", "unknown"


def inspect_file(path: Path) -> dict[str, Any]:
    """Return column names, counts, preview, and light validation hints."""
    matrix, err, kind = _load_matrix(path)
    if err:
        return {
            "ok": False,
            "format": kind,
            "error": err,
            "columns": [],
            "rowCount": 0,
            "dataRowCount": 0,
            "preview": [],
            "hints": [],
        }

    if not matrix:
        return {
            "ok": False,
            "format": kind,
            "error": "File appears empty",
            "columns": [],
            "rowCount": 0,
            "dataRowCount": 0,
            "preview": [],
            "hints": [],
        }

    header = matrix[0]
    data_rows = matrix[1:]
    col_count = max(len(header), max((len(r) for r in data_rows), default=0))

    def pad(row: list[str]) -> list[str]:
        r = list(row) + [""] * max(0, col_count - len(row))
        return r[:col_count]

    padded_header = pad(header)
    preview = [pad(r) for r in data_rows[:_PREVIEW_ROWS]]

    hints: list[str] = []
    if col_count == 0:
        hints.append("No columns detected.")
    if not any(any(c for c in r) for r in data_rows):
        hints.append("No data rows after header.")

    return {
        "ok": True,
        "format": kind,
        "columns": padded_header,
        "rowCount": len(matrix),
        "dataRowCount": len(data_rows),
        "preview": [padded_header] + preview if preview else [padded_header],
        "hints": hints,
    }


def clean_file(path: Path, options: dict[str, Any] | None = None) -> dict[str, Any]:
    """Strip cells, drop blank rows, normalize headers; return previews and stats."""
    matrix, err, kind = _load_matrix(path)
    if err:
        return {"ok": False, "format": kind, "error": err}

    if not matrix:
        return {"ok": False, "format": kind, "error": "Nothing to clean"}

    orig_data_count = max(0, len(matrix) - 1)
    col_width = max(
        len(matrix[0]),
        max((len(r) for r in matrix[1:]), default=0),
    )

    def pad_row(r: list[str]) -> list[str]:
        rr = [_cell_str(c) for c in r] + [""] * max(0, col_width - len(r))
        return rr[:col_width]

    padded = [pad_row(r) for r in matrix]
    header_cells = padded[0]
    norm_heads = [_norm_header(h) for h in header_cells]

    seen: dict[str, int] = {}
    final_heads: list[str] = []
    header_changes: list[dict[str, str]] = []
    for i, nh in enumerate(norm_heads):
        base = nh or f"column_{i+1}"
        cnt = seen.get(base, 0)
        seen[base] = cnt + 1
        fn = base if cnt == 0 else f"{base}_{cnt + 1}"
        final_heads.append(fn)
        orig = header_cells[i] if i < len(header_cells) else ""
        if orig != fn:
            header_changes.append({"from": orig or "(empty)", "to": fn})

    opts = options or {}
    remove_duplicates = bool(opts.get("removeDuplicates", True))
    fill_numeric = bool(opts.get("fillNumeric", True))
    fill_categorical = bool(opts.get("fillCategorical", True))
    standardize_text = bool(opts.get("standardizeText", False))

    cleaned: list[list[str]] = [final_heads]
    removed_empty = 0
    for r in padded[1:]:
        row = pad_row(r)[: len(final_heads)]
        while len(row) < len(final_heads):
            row.append("")
        row = row[: len(final_heads)]
        if not any(row):
            removed_empty += 1
            continue
        cleaned.append(row)

    body = cleaned[1:]

    duplicates_removed = 0
    if remove_duplicates and body:
        seen: set[str] = set()
        deduped: list[list[str]] = []
        for row in body:
            key = "|".join(row)
            if key in seen:
                duplicates_removed += 1
                continue
            seen.add(key)
            deduped.append(row)
        body = deduped

    inferred_numeric: list[int] = []
    inferred_categorical: list[int] = []
    column_missing_map: dict[int, int] = {}
    missing_values_before_fill = 0
    for ci in range(len(final_heads)):
        vals = [r[ci] for r in body if ci < len(r)]
        missing_count = sum(1 for v in vals if _is_missing(v))
        column_missing_map[ci] = missing_count
        missing_values_before_fill += missing_count
        present = [v for v in vals if not _is_missing(v)]
        all_num = bool(present) and all(_to_float(v) is not None for v in present)
        if all_num:
            inferred_numeric.append(ci)
        else:
            inferred_categorical.append(ci)

    column_summary: list[dict[str, Any]] = []
    numeric_idx = set(inferred_numeric)
    for ci, name in enumerate(final_heads):
        is_num = ci in numeric_idx
        column_summary.append(
            {
                "name": name,
                "type": "number" if is_num else "string",
                "category": "Quantitative" if is_num else "Qualitative",
                "missing": int(column_missing_map.get(ci, 0)),
            }
        )

    numeric_fill_map: dict[int, str] = {}
    if fill_numeric:
        for ci in inferred_numeric:
            nums = [_to_float(r[ci]) for r in body if ci < len(r)]
            nums = [n for n in nums if n is not None]
            med = _median(nums)
            numeric_fill_map[ci] = str(int(med)) if float(med).is_integer() else str(round(med, 6))

    categorical_fill_map: dict[int, str] = {}
    if fill_categorical:
        for ci in inferred_categorical:
            vals = [r[ci] for r in body if ci < len(r) and not _is_missing(r[ci])]
            freq: dict[str, int] = {}
            for v in vals:
                freq[v] = freq.get(v, 0) + 1
            if not freq:
                categorical_fill_map[ci] = "Unknown"
            else:
                categorical_fill_map[ci] = max(freq.items(), key=lambda x: x[1])[0]

    standardized_cells = 0
    numeric_filled_cells = 0
    categorical_filled_cells = 0
    final_body: list[list[str]] = []
    for row in body:
        rr = list(row)
        for ci in range(len(final_heads)):
            if ci >= len(rr):
                rr.append("")
            cell = rr[ci]
            if standardize_text and isinstance(cell, str):
                std = re.sub(r"\s+", " ", cell.strip())
                if std != cell:
                    standardized_cells += 1
                rr[ci] = std
            if _is_missing(rr[ci]) and fill_numeric and ci in numeric_fill_map:
                rr[ci] = numeric_fill_map[ci]
                numeric_filled_cells += 1
            elif _is_missing(rr[ci]) and fill_categorical and ci in categorical_fill_map:
                rr[ci] = categorical_fill_map[ci]
                categorical_filled_cells += 1
        final_body.append(rr[: len(final_heads)])

    cleaned = [final_heads] + final_body

    orig_preview = padded[: 1 + min(_PREVIEW_ROWS, max(0, len(padded) - 1))]
    clean_preview = cleaned[: 1 + min(_PREVIEW_ROWS, max(0, len(cleaned) - 1))]
    valuation_hints = summarize_for_valuation(cleaned)

    return {
        "ok": True,
        "format": kind,
        "success": True,
        "originalPreview": orig_preview,
        "cleanedPreview": clean_preview,
        "changes": {
            "originalDataRows": orig_data_count,
            "cleanedDataRows": max(0, len(cleaned) - 1),
            "removedEmptyRows": removed_empty,
            "duplicatesRemoved": duplicates_removed,
            "missingValuesBeforeFill": missing_values_before_fill,
            "numericColumnsDetected": len(inferred_numeric),
            "categoricalColumnsDetected": len(inferred_categorical),
            "numericCellsFilled": numeric_filled_cells,
            "categoricalCellsFilled": categorical_filled_cells,
            "textCellsStandardized": standardized_cells,
            "headerNormalizations": header_changes[:30],
            "columnSummary": column_summary,
            "appliedOptions": {
                "removeDuplicates": remove_duplicates,
                "fillNumeric": fill_numeric,
                "fillCategorical": fill_categorical,
                "standardizeText": standardize_text,
            },
        },
        "columnsCleaned": final_heads,
        "valuationHints": valuation_hints,
    }


def summarize_for_valuation(cleaned_rows: list[list[str]]) -> dict[str, Any]:
    """Lightweight aggregates on cleaned tabular data (headers + body)."""
    if len(cleaned_rows) < 2:
        return {
            "rowCount": 0,
            "numericColumnSamples": [],
            "note": "No data rows after cleaning.",
        }

    headers = cleaned_rows[0]
    body = cleaned_rows[1:]
    numeric_keywords = (
        "value",
        "amount",
        "face",
        "price",
        "principal",
        "notional",
        "balance",
        "yield",
        "rate",
        "coupon",
    )

    samples: list[dict[str, Any]] = []
    for j, name in enumerate(headers):
        key = name.lower()
        if not any(k in key for k in numeric_keywords):
            continue
        vals: list[float] = []
        for r in body:
            if j >= len(r):
                continue
            raw = (r[j] or "").replace(",", "")
            try:
                vals.append(float(raw))
            except ValueError:
                continue
        if not vals:
            continue
        samples.append(
            {
                "column": name,
                "count": len(vals),
                "sum": round(sum(vals), 4),
                "min": round(min(vals), 4),
                "max": round(max(vals), 4),
            }
        )

    return {
        "rowCount": len(body),
        "numericColumnSamples": samples[:12],
        "note": "Figures are parsed from cleaned upload data for QA; DB instruments remain authoritative.",
    }
